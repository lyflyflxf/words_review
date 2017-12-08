#!/sr/bin/env python
# -*- coding: gbk -*-

# inter=[0,1,2,4,7,15,30]
# for i in range(1, len(inter)):
#     day_list.append(sum(inter[:i]))
# print(day_list)

import openpyxl
from datetime import *
import os
import pyperclip

interval = [1, 2, 4, 7, 15, 30]
review_day = [1, 3, 7, 14, 29, 59]

f_dir = r"E:\py\tools\s_plan\\"
f_name = 't.xlsx'
wb = openpyxl.load_workbook(f_dir + f_name)


# wb.template=False

class Record:
    GENERAL = 'General'
    DATE = 'yyyy/m/d'

    def r_value(self,column):
        return sheet.cell(None, self.row, column).value

    def w_value(self,column, value, format=GENERAL):
        cell = sheet.cell(None, self.row, column)
        cell.value = value
        cell.number_format = format

    # def update(self):


    def __init__(self, name, row):
        self.name = name
        self.row = row
        sheet = wb.get_sheet_by_name(name)

        def future(sn):
            if not (isinstance(sn, int) and sn >= 0):
                raise ValueError('��ϰ������Ų�����Ȼ��')
            l = len(interval)
            if sn < l:
                return interval[sn]
            else:
                return 60

        self._init, self._content, self._sn, self._last, self._next, self._dl = \
            (self.r_value(i) for i in range(1, 7))
        # self.init=r_value(1)
        # self.content=r_value(2)
        # self.sn=r_value(3)
        # self.last=r_value(4)
        # self.next=r_value(5)
        # self.dl=r_value(6)

        if self._sn == None:
            self._sn = 0
            self.w_value(3, self._sn)
        if self._last == None:
            self._last = self._init
            self.w_value(4, self._last, self.DATE)
        if self._next == None:
            self._next = self._last + timedelta(future(self._sn))
            self.w_value(5, self._next, self.DATE)
        if self._dl == None:
            self._dl = self._next + timedelta(future(self._sn + 1))
        #     w_value(6, self.dl, DATE)

    def __repr__(self):
        return self._content

    def verify(self, date):
        # elapse = (date - self.last).days  # ���� - �ϴθ�ϰ����
        # if elapse == interval[self.sn] or elapse % 30 == 0:
        if date >= self._next and date <= self._dl:
            return True
        else:
            return False

    @property
    def today(self):
        return self.verify(datetime.now())

    @property
    def tomorrow(self):
        return self.verify(datetime.now() + timedelta(1))


students = wb.get_sheet_names()

print('������', date.today())

for name in students:
    sheet = wb.get_sheet_by_name(name)

    tasks = {'����': [], '����': []}
    today = []
    tomorrow = []

    i = 2
    while i < sheet.max_row:
        if sheet.cell(row=i, column=2).value is None:
            tasks.update({'����': today})
            tasks.update({'����': tomorrow})
            break
        else:
            r = Record(name, i)
            if r.today:  # ��ȡ������������
                today.append(r)
            elif r.tomorrow:  # ��ȡ������������
                tomorrow.append(r)
        i += 1

    print('ѧ�� ', name)
    for k, v in tasks.items():
        out=''
        if len(v) == 0:
            print(k + '�޸�ϰ����')
        else:
            print(k + '�ĸ�ϰ����Ϊ��')
        for item in v:
            out+=item._content+' '
        print(out)
        pyperclip.copy(out)

# wb.save(f_dir + f_name)

# os.startfile(f_dir + f_name)